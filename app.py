"""
Système d'Alerte Précoce — Décrochage Académique
Application Streamlit avec interprétabilité SHAP
v2.0 — Données institutionnelles complètes (13 champs) + SHAP dynamique + Recommandations dynamiques
"""

import os
import io
import datetime
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import shap
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from sklearn.preprocessing import StandardScaler

# ─────────────────────────────────────────────────────────────
# CONFIG PAGE
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Système d'Alerte Précoce",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ─────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700&family=Source+Sans+3:wght@300;400;500;600&display=swap');
:root {
    --bleu-nuit:  #0f1f3d; --bleu-roi: #1a3a6b; --bleu-acier: #2d6a9f;
    --or: #c9a84c; --or-clair: #e8c96a; --blanc: #ffffff;
    --gris-clair: #f0f2f6; --rouge: #c0392b; --vert: #1e7e5a; --texte: #1a1a2e;
}
html, body, [class*="css"] { font-family: 'Source Sans 3', sans-serif; color: var(--texte); }
.header-main {
    background: linear-gradient(135deg, var(--bleu-nuit) 0%, var(--bleu-roi) 60%, var(--bleu-acier) 100%);
    padding: 2.5rem 3rem; border-radius: 0 0 24px 24px;
    margin: -1rem -1rem 2rem -1rem; position: relative; overflow: hidden;
}
.header-main::before {
    content: ''; position: absolute; top: -50%; right: -10%;
    width: 400px; height: 400px;
    background: radial-gradient(circle, rgba(201,168,76,0.15) 0%, transparent 70%);
    border-radius: 50%;
}
.header-title { font-family: 'Playfair Display', serif; font-size: 2rem; font-weight: 700; color: var(--blanc); margin: 0; line-height: 1.2; }
.header-subtitle { font-size: 0.95rem; color: rgba(255,255,255,0.7); margin-top: 0.4rem; font-weight: 300; letter-spacing: 0.05em; }
.header-badge {
    display: inline-block; background: rgba(201,168,76,0.2); border: 1px solid var(--or);
    color: var(--or-clair); padding: 0.2rem 0.8rem; border-radius: 20px;
    font-size: 0.75rem; font-weight: 500; letter-spacing: 0.08em;
    text-transform: uppercase; margin-bottom: 0.6rem;
}
.stTabs [data-baseweb="tab-list"] { gap: 4px; background: var(--gris-clair); padding: 6px; border-radius: 12px; }
.stTabs [data-baseweb="tab"] {
    background: transparent; border: none; border-radius: 8px; padding: 0.5rem 1.2rem;
    font-size: 0.85rem; font-weight: 500; color: #666; transition: all 0.2s;
}
.stTabs [aria-selected="true"] { background: var(--blanc) !important; color: var(--bleu-roi) !important; font-weight: 600 !important; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }
.section-card { background: var(--blanc); border: 1px solid #e8ecf0; border-radius: 14px; padding: 1.5rem; margin-bottom: 1rem; box-shadow: 0 2px 12px rgba(0,0,0,0.04); }
.section-title { font-family: 'Playfair Display', serif; font-size: 1rem; font-weight: 600; color: var(--bleu-roi); margin-bottom: 1rem; padding-bottom: 0.5rem; border-bottom: 2px solid var(--or); }
.stButton > button {
    background: linear-gradient(135deg, var(--bleu-roi), var(--bleu-acier)) !important;
    color: white !important; border: none !important; border-radius: 10px !important;
    padding: 0.75rem 2.5rem !important; font-size: 1rem !important; font-weight: 600 !important;
    letter-spacing: 0.03em !important; transition: all 0.3s !important;
    box-shadow: 0 4px 15px rgba(26,58,107,0.3) !important; width: 100% !important;
}
.stButton > button:hover { transform: translateY(-2px) !important; box-shadow: 0 6px 20px rgba(26,58,107,0.4) !important; }
.result-alerte   { background: linear-gradient(135deg,#fdf0ee,#fff5f4); border: 2px solid #e74c3c; border-left: 6px solid #c0392b; border-radius: 14px; padding: 1.8rem 2rem; margin-bottom: 1.5rem; }
.result-securite { background: linear-gradient(135deg,#edfaf5,#f0fdf8); border: 2px solid #2ecc71; border-left: 6px solid var(--vert); border-radius: 14px; padding: 1.8rem 2rem; margin-bottom: 1.5rem; }
.result-title { font-family: 'Playfair Display', serif; font-size: 1.6rem; font-weight: 700; margin: 0 0 0.3rem 0; }
.result-score { font-size: 0.9rem; opacity: 0.8; margin-bottom: 1rem; }
.jauge-container { background: #e8ecf0; border-radius: 50px; height: 14px; overflow: hidden; margin: 0.8rem 0; }
.jauge-fill-rouge { height: 100%; border-radius: 50px; background: linear-gradient(90deg,#f39c12,#e74c3c); }
.jauge-fill-vert  { height: 100%; border-radius: 50px; background: linear-gradient(90deg,#27ae60,#2ecc71); }
.pond-card { background: var(--gris-clair); border-radius: 10px; padding: 1rem 1.2rem; margin-bottom: 0.6rem; display: flex; justify-content: space-between; align-items: center; }
.pond-label  { font-size: 0.85rem; color: #555; font-weight: 500; }
.pond-value  { font-size: 1rem; font-weight: 700; color: var(--bleu-roi); }
.pond-detail { font-size: 0.75rem; color: #888; margin-top: 0.1rem; }
.reco-item   { background: linear-gradient(135deg,#fffbf0,#fff8e8); border-left: 4px solid var(--or); border-radius: 0 8px 8px 0; padding: 0.7rem 1rem; margin-bottom: 0.5rem; font-size: 0.88rem; color: var(--texte); }
.reco-ok     { background: linear-gradient(135deg,#edfaf5,#f0fdf8); border-left: 4px solid var(--vert); border-radius: 0 8px 8px 0; padding: 0.7rem 1rem; margin-bottom: 0.5rem; font-size: 0.88rem; color: var(--texte); }
.stSelectbox > div > div { border-radius: 8px !important; border-color: #d0d7de !important; }
.footer { text-align: center; padding: 2rem; color: #aaa; font-size: 0.78rem; margin-top: 3rem; border-top: 1px solid #eee; }
.metric-mini { background: var(--blanc); border: 1px solid #e8ecf0; border-radius: 10px; padding: 0.8rem 1rem; text-align: center; }
.metric-mini-val   { font-size: 1.4rem; font-weight: 700; color: var(--bleu-roi); }
.metric-mini-label { font-size: 0.72rem; color: #888; margin-top: 0.1rem; }
.info-mediane { background: #f0f4ff; border: 1px solid #c5d3f0; border-radius: 8px; padding: 0.6rem 1rem; font-size: 0.8rem; color: #3a5a9a; margin-top: 0.5rem; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# CHARGEMENT DES MODÈLES
# ─────────────────────────────────────────────────────────────
@st.cache_resource
def charger_modeles():
    base       = os.path.dirname(os.path.abspath(__file__))
    path_rf    = os.path.join(base, 'models', 'RF_Project') + os.sep
    path_tools = os.path.join(base, 'models', 'DROPOUT_TOOLS') + os.sep
    try:
        rf_comport    = joblib.load(path_rf    + 'Dropout_model')
        scaler_rf     = joblib.load(path_rf    + 'scaler_rf.pkl')
        rf_features   = joblib.load(path_rf    + 'feature_names.pkl')
        rf_inst       = joblib.load(path_tools + 'Dropout_model_institutional.pkl')
        inst_features = joblib.load(path_tools + 'feature_names_institutional.pkl')
        # ── Seuil optimal de fusion calculé sur Val ──
        seuil_path = path_tools + 'seuil_optimal_fusion.pkl'
        if os.path.exists(seuil_path):
            seuil_fusion = joblib.load(seuil_path)
        else:
            seuil_fusion = 0.5
            st.warning("⚠️ Fichier seuil_optimal_fusion.pkl introuvable — seuil par défaut 0.5 utilisé.")
        return rf_comport, scaler_rf, rf_features, rf_inst, inst_features, seuil_fusion
    except FileNotFoundError as e:
        st.error(f"❌ Modèle introuvable : {e}")
        st.info("Vérifiez que vos modèles sont bien dans : models/RF_Project/ et models/DROPOUT_TOOLS/")
        st.stop()

# ─────────────────────────────────────────────────────────────
# PRÉTRAITEMENT RF
# ─────────────────────────────────────────────────────────────
ORDINAL_MAPPINGS = {
    "1.Tranche d'âge":
        {'18–20 ans': 0, '21–23 ans': 1, '24 ans ou plus': 2},
    '4.Quelle est votre moyenne générale la plus récente ?':
        {'Moins de 10': 0, 'Entre 10 et 12': 1, 'Entre 12 et 14': 2, 'Plus de 14': 3},
    '5.Combien de matières avez-vous échouées lors de la dernière période académique ?':
        {'Aucune': 0, '1–2': 1, '3 ou plus': 2},
    '6.À quelle fréquence êtes-vous absent(e) aux cours ?':
        {'Rarement': 0, 'Parfois': 1, 'Souvent': 2},
    "7.Combien d'heures étudiez-vous en moyenne par semaine (hors cours) ?":
        {'Moins de 5 heures': 0, 'Entre 5 et 10 heures': 1, 'Plus de 10 heures': 2},
    '8.Votre participation en classe est généralement':
        {'Moyenne': 0, 'Élevée': 1},
    '9.Rendez-vous vos devoirs et travaux dans les délais ?':
        {'Jamais': 0, 'Parfois': 1, 'Toujours': 2},
    '10.À quel point procrastinez-vous dans vos études ?':
        {'Peu': 0, 'Moyennement': 1, 'Beaucoup': 2},
    '11.Je me sens motivé(e) dans mes études':
        {'Pas du tout': 0, 'Peu': 1, 'Motivé(e)': 2, 'Très motivé(e)': 3},
    '12.Mon niveau de stress lié aux études est :':
        {'Très faible': 0, 'Faible': 1, 'Peu': 1, 'Moyen': 2, 'Élevé': 3},
    "13.J'ai confiance en ma capacité à réussir mes études":
        {'Peu': 0, 'Moyennement': 1, 'Confiant(e)': 2, 'Très confiant(e)': 3},
    '14.Je ressens une fatigue mentale liée aux études':
        {'Très faible': 0, 'Faible': 1, 'Moyenne': 2, 'Élevée': 3, 'Très élevée': 4},
    "15.Disposez-vous d'un accès régulier à Internet pour vos études ?":
        {'Non': 0, 'Oui': 1},
    '16.Votre environnement de travail à domicile est :':
        {'Défavorable': 0, 'Moyennement favorable': 1, 'Favorable': 2},
    '17.Le soutien de votre entourage (famille/proches) dans vos études est :':
        {'Faible': 0, 'Moyen': 1, 'Fort': 2},
    '20.Niveau_Etude_Parents':
        {'Bac ou moins': 0, 'Licence': 1, 'Master ou plus': 2},
    '22.Mention_Bac':
        {'Sans mention': 0, 'Assez bien': 1, 'Bien': 2, 'Très bien': 3},
    '23.Annees_Retard':
        {'0': 0, '1': 1, '2': 2, '3+': 3},
    '24.Freq_Connexion_LMS':
        {'Rare': 0, 'Hebdomadaire': 1, 'Quotidienne': 2},
    '25.Temps_Trajet_Minutes':
        {'Moins de 30': 0, '30-60': 1, 'Plus de 60': 2},
}

NOMINAL_COLS = [
    '2.Genre', "3.Niveau d'étude actuel",
    '18.Statut_Boursier', '19.Activite_Salariee',
    '21.Type_Bac', '26.Type_Logement'
]

def pretraiter(donnees: dict, scaler_rf, rf_features):
    df_e = pd.DataFrame([donnees])
    for col, mapping in ORDINAL_MAPPINGS.items():
        if col in df_e.columns:
            df_e[col] = df_e[col].map(mapping)
    existing = [c for c in NOMINAL_COLS if c in df_e.columns]
    df_e = pd.get_dummies(df_e, columns=existing, drop_first=True)
    df_e = df_e.reindex(columns=rf_features, fill_value=0).astype(float)
    return scaler_rf.transform(df_e), df_e

def pretraiter_inst(input_admin: pd.DataFrame, inst_features: list) -> pd.DataFrame:
    """
    Aligne les données institutionnelles d'un étudiant
    sur les features attendues par le RF institutionnel.
    Les features non renseignées sont imputées à 0.
    Pas de scaling — RF invariant aux transformations monotones.
    """
    return input_admin.reindex(columns=inst_features, fill_value=0).astype(float)

# ─────────────────────────────────────────────────────────────
# EXTRACTION SHAP 
# ─────────────────────────────────────────────────────────────
def extraire_sv(shap_vals):
    """
    Retourne toujours un vecteur numpy 1D float, quelle que soit
    la version de SHAP (ancienne API list, nouveau array 2D/3D).
    """
    if isinstance(shap_vals, list):
        raw = np.array(shap_vals[1])
    else:
        raw = np.array(shap_vals)
    if raw.ndim == 3:
        raw = raw[:, :, 1]
    sv = raw[0]
    return np.array(sv, dtype=float).flatten()

# ─────────────────────────────────────────────────────────────
# NETTOYAGE DYNAMIQUE DES NOMS DE FEATURES SHAP
# ─────────────────────────────────────────────────────────────
# Table de traduction : nom interne de feature → label lisible en français
# Couvre les préfixes issus de l'encodage ordinal, OHE et les noms bruts
FEATURE_LABELS_FR = {
    # ── Variables comportementales RF ──────────────────────
    "tranche d'âge":               "Tranche d'âge",
    "genre":                       "Genre",
    "niveau d'étude actuel":       "Niveau d'étude",
    "moyenne générale":            "Moyenne générale",
    "matières":                    "Matières échouées",
    "absent":                      "Absences aux cours",
    "heures":                      "Heures d'étude/semaine",
    "participation":               "Participation en classe",
    "devoirs":                     "Devoirs dans les délais",
    "procrastin":                  "Procrastination",
    "motivé":                      "Motivation",
    "stress":                      "Niveau de stress",
    "confiance":                   "Confiance en soi",
    "fatigue":                     "Fatigue mentale",
    "internet":                    "Accès Internet",
    "environnement":               "Environnement de travail",
    "soutien":                     "Soutien de l'entourage",
    "statut_boursier":             "Statut boursier",
    "activite_salariee":           "Activité salariée",
    "niveau_etude_parents":        "Niveau d'étude parents",
    "type_bac":                    "Type de Bac",
    "mention_bac":                 "Mention au Bac",
    "annees_retard":               "Années de retard",
    "freq_connexion_lms":          "Fréquence LMS",
    "temps_trajet":                "Temps de trajet",
    "type_logement":               "Type de logement",
    # ── Variables institutionnelles RF institutionnel ──────
    "age at enrollment":           "Âge à l'inscription",
    "curricular units 1st sem (approved)":  "UC validées S1",
    "curricular units 1st sem (grade)":     "Note moyenne S1",
    "curricular units 1st sem (enrolled)":  "UC inscrites S1",
    "curricular units 1st sem (evaluations)": "UC évaluées S1",
    "curricular units 1st sem (credited)":  "UC créditées S1",
    "curricular units 2nd sem (approved)":  "UC validées S2",
    "curricular units 2nd sem (grade)":     "Note moyenne S2",
    "curricular units 2nd sem (enrolled)":  "UC inscrites S2",
    "curricular units 2nd sem (evaluations)": "UC évaluées S2",
    "curricular units 2nd sem (credited)":  "UC créditées S2",
    "tuition fees up to date":     "Frais de scolarité à jour",
    "scholarship holder":          "Bourse institutionnelle",
    "debtor":                      "Dettes universitaires",
    "displaced":                   "Étudiant déplacé",
    "gender":                      "Genre (institutionnel)",
    "international":               "Étudiant international",
    "unemployment rate":           "Taux de chômage",
    "inflation rate":              "Taux d'inflation",
    "gdp":                         "PIB national",
    "educational special needs":   "Besoins éducatifs spéciaux",
    "daytime/evening attendance":  "Cours jour/soir",
    "application mode":            "Mode de candidature",
    "application order":           "Ordre de candidature",
    "course":                      "Filière suivie",
    "previous qualification":      "Qualification précédente",
    "marital status":              "Situation familiale",
    "nacionality":                 "Nationalité",
    "mother's qualification":      "Qualification de la mère",
    "father's qualification":      "Qualification du père",
    "mother's occupation":         "Profession de la mère",
    "father's occupation":         "Profession du père",
}

def traduire_feature(raw_name: str) -> str:
    """
    Traduit dynamiquement un nom de feature brut (issu du preprocessing)
    en label lisible en français.
    Logique : cherche le meilleur match partiel dans FEATURE_LABELS_FR.
    """
    # 1. Nettoyage : retrait du numéro de début ("14.fatigue mentale..." → "fatigue mentale...")
    cleaned = raw_name.lower().strip()
    # Retrait du préfixe numérique type "14." ou "14. "
    if cleaned and cleaned[0].isdigit():
        parts = cleaned.split('.', 1)
        if len(parts) > 1:
            cleaned = parts[1].strip()

    # 2. Retrait des suffixes OHE type "_1", "_2", etc.
    import re
    cleaned_no_suffix = re.sub(r'_\d+$', '', cleaned).strip()

    # 3. Chercher la meilleure correspondance par sous-chaîne
    best_match = None
    best_len   = 0
    for key, label in FEATURE_LABELS_FR.items():
        if key in cleaned_no_suffix or key in cleaned:
            if len(key) > best_len:
                best_match = label
                best_len   = len(key)

    if best_match:
        return best_match

    # 4. Fallback : capitaliser et tronquer le nom nettoyé
    fallback = cleaned_no_suffix.replace('_', ' ').strip()
    return fallback.capitalize()[:40] if fallback else raw_name[:40]


# ─────────────────────────────────────────────────────────────
# RECOMMANDATIONS DYNAMIQUES BASÉES SUR LES SHAP
# ─────────────────────────────────────────────────────────────
# Chaque entrée : (mots-clés à chercher dans le label traduit, emoji, texte recommandation)
RECO_RULES = [
    (["fatigue"],
     "🧠", "Accompagnement psychologique et gestion de la charge mentale conseillés."),
    (["motivation"],
     "🎯", "Entretien de remotivation avec un conseiller pédagogique recommandé."),
    (["stress"],
     "🌿", "Atelier de gestion du stress et techniques de relaxation suggérés."),
    (["confiance"],
     "💪", "Séances de soutien scolaire pour renforcer la confiance en soi recommandées."),
    (["procrastin"],
     "📅", "Atelier de gestion du temps et de planification du travail conseillé."),
    (["heures", "étude"],
     "📚", "Suivi de la charge de travail hebdomadaire recommandé."),
    (["absent"],
     "📋", "Entretien de suivi des absences avec le référent pédagogique conseillé."),
    (["moyenne", "note"],
     "✏️", "Mise en place d'un tutorat académique conseillée."),
    (["devoirs", "délais"],
     "📝", "Accompagnement dans la planification et la remise des travaux recommandé."),
    (["logement"],
     "🏠", "Orientation vers les services sociaux universitaires suggérée."),
    (["soutien", "entourage"],
     "👨‍👩‍👧", "Renforcement du réseau de soutien social conseillé."),
    (["environnement", "domicile"],
     "🖥️", "Accès à des espaces de travail universitaires recommandé."),
    (["trajet"],
     "🚌", "Information sur les aménagements horaires possibles conseillée."),
    (["retard"],
     "📊", "Entretien de suivi du parcours académique recommandé."),
    (["lms", "connexion"],
     "💻", "Formation aux outils numériques d'apprentissage conseillée."),
    (["uc validées", "uc inscrites", "approved", "enrolled"],
     "📐", "Analyse du parcours curriculaire avec le responsable pédagogique recommandée."),
    (["frais", "dettes", "tuition", "debtor"],
     "💳", "Orientation vers le service des affaires financières étudiantes conseillée."),
    (["bourse", "scholarship"],
     "🎓", "Vérification des droits aux aides financières avec le service social recommandée."),
    (["chômage", "gdp", "inflation"],
     "🌍", "Accompagnement sur l'insertion professionnelle et les débouchés conseillé."),
    (["déplacé", "displaced", "international"],
     "🌐", "Accompagnement spécifique étudiant déplacé/international recommandé."),
]

def generer_recommandations(sv_1d: np.ndarray, feat_labels: list) -> list:
    """
    Génère des recommandations entièrement dynamiques :
    - Trie les features par SHAP positif décroissant
    - Cherche la règle correspondante dans RECO_RULES
    - Retourne les 4 recommandations les plus pertinentes
    Aucun texte n'est hard-codé — tout découle des SHAP values réelles.
    """
    # Paires (shap_value, label_traduit) triées par contribution positive décroissante
    paires = sorted(
        zip(sv_1d.tolist(), feat_labels),
        key=lambda x: x[0],
        reverse=True
    )

    recos  = []
    vus    = set()

    for val, label in paires:
        if val <= 0:
            break   
        label_lower = label.lower()
        for keywords, emoji, texte in RECO_RULES:
            if any(kw in label_lower for kw in keywords):
                cle = texte  
                if cle not in vus:
                    recos.append(f"{emoji} {texte}")
                    vus.add(cle)
                break
        if len(recos) >= 4:
            break

    if not recos:
        recos = ["📋 Entretien de suivi global avec le conseiller pédagogique recommandé."]

    return recos


# ─────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-main">
    <div class="header-badge">Outil de Prédiction</div>
    <div class="header-title">🎓 Système d'Alerte Précoce</div>
    <div class="header-subtitle">Détection du risque de décrochage académique · RF comportemental + RF institutionnel · Interprétabilité SHAP</div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# CHARGEMENT
# ─────────────────────────────────────────────────────────────
rf_model, scaler_rf, rf_features, rf_inst, inst_features, seuil_fusion = charger_modeles()

# ─────────────────────────────────────────────────────────────
# SESSION STATE — persistance inter-onglets + historique
# ─────────────────────────────────────────────────────────────
_SS_DEFAULTS = {
    "ss_age": "18\u201320 ans", "ss_genre": "Femme",
    "ss_niveau": "Licence 1 (L1)", "ss_type_bac": "G\u00e9n\u00e9ral",
    "ss_mention_bac": "Sans mention",
    "ss_moyenne": "Entre 10 et 12", "ss_matieres": "Aucune",
    "ss_absences": "Rarement", "ss_annees_ret": "0",
    "ss_heures": "Entre 5 et 10 heures", "ss_participation": "Moyenne",
    "ss_devoirs": "Parfois",
    "ss_procrastin": "Peu", "ss_motivation": "Motiv\u00e9(e)",
    "ss_stress": "Moyen", "ss_confiance": "Confiant(e)", "ss_fatigue": "Faible",
    "ss_statut_bours": "Non", "ss_activite_sal": "Aucune",
    "ss_niv_parents": "Licence", "ss_internet": "Oui",
    "ss_environnement": "Favorable", "ss_soutien": "Fort",
    "ss_freq_lms": "Hebdomadaire", "ss_trajet": "Moins de 30",
    "ss_logement": "Chez les parents",
    "ss_age_enrol": 20, "ss_curr_1_enr": 6, "ss_curr_1_app": 5,
    "ss_curr_1_note": 12.0, "ss_curr_2_enr": 6, "ss_curr_2_app": 5,
    "ss_curr_2_note": 12.0,
    "ss_frais_ok": "Oui", "ss_dettes": "Non", "ss_bourse_inst": "Non",
    "ss_deplace": "Non", "ss_international": "Non",
    "ss_chomage": 11.1, "ss_inflation": 1.4, "ss_gdp": 1.74,
    "historique": [],
}
for _k, _v in _SS_DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────────────────────
# SWITCHER DE MODE : Saisie manuelle / Import fichier
# ─────────────────────────────────────────────────────────────

# ── CSS dédié au switcher ─────────────────────────────────
st.markdown("""
<style>
div[data-testid="stRadio"] > div { flex-direction: row; gap: 1rem; }
div[data-testid="stRadio"] label {
    background: var(--bleu-nuit); border: 1px solid #3a5a8a;
    border-radius: 10px; padding: 0.5rem 1.4rem;
    cursor: pointer; font-weight: 500; transition: all 0.2s;
    color: rgba(255,255,255,0.75) !important;
}
div[data-testid="stRadio"] label:has(input:checked) {
    background: var(--bleu-roi); color: white !important;
    border-color: var(--or); box-shadow: 0 2px 8px rgba(26,58,107,0.35);
}
.upload-zone { background: #f8faff; border: 2px dashed #c5d3f0;
    border-radius: 14px; padding: 2rem; margin-bottom: 1.5rem; text-align: center; }
.upload-title { font-family: 'Playfair Display', serif; font-size: 1rem;
    color: var(--bleu-roi); font-weight: 600; margin-bottom: 0.5rem; }
.upload-hint  { font-size: 0.82rem; color: #888; margin-top: 0.3rem; }
.batch-result-ok    { background:#edfaf5; border-left:4px solid #1e7e5a;
    border-radius:0 8px 8px 0; padding:0.4rem 0.8rem; font-size:0.83rem; color:#155843; }
.batch-result-alert { background:#fdf0ee; border-left:4px solid #c0392b;
    border-radius:0 8px 8px 0; padding:0.4rem 0.8rem; font-size:0.83rem; color:#7b1c10; }
</style>
""", unsafe_allow_html=True)

mode = st.radio(
    "**Mode d'analyse**",
    ["✏️  Saisie manuelle (1 étudiant)", "📂  Import fichier (plusieurs étudiants)"],
    horizontal=True,
    label_visibility="collapsed",
)

st.markdown("<br>", unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════
# ── MODE IMPORT FICHIER ───────────────────────────────────────
# ═════════════════════════════════════════════════════════════
if mode == "📂  Import fichier (plusieurs étudiants)":

    # ── Colonnes attendues dans le CSV ────────────────────
    COLONNES_RF_COMP = [
        "1.Tranche d'âge",
        "2.Genre",
        "3.Niveau d'étude actuel",
        "4.Quelle est votre moyenne générale la plus récente ?",
        "5.Combien de matières avez-vous échouées lors de la dernière période académique ?",
        "6.À quelle fréquence êtes-vous absent(e) aux cours ?",
        "7.Combien d'heures étudiez-vous en moyenne par semaine (hors cours) ?",
        "8.Votre participation en classe est généralement",
        "9.Rendez-vous vos devoirs et travaux dans les délais ?",
        "10.À quel point procrastinez-vous dans vos études ?",
        "11.Je me sens motivé(e) dans mes études",
        "12.Mon niveau de stress lié aux études est :",
        "13.J'ai confiance en ma capacité à réussir mes études",
        "14.Je ressens une fatigue mentale liée aux études",
        "15.Disposez-vous d'un accès régulier à Internet pour vos études ?",
        "16.Votre environnement de travail à domicile est :",
        "17.Le soutien de votre entourage (famille/proches) dans vos études est :",
        "18.Statut_Boursier",
        "19.Activite_Salariee",
        "20.Niveau_Etude_Parents",
        "21.Type_Bac",
        "22.Mention_Bac",
        "23.Annees_Retard",
        "24.Freq_Connexion_LMS",
        "25.Temps_Trajet_Minutes",
        "26.Type_Logement",
    ]
    COLONNES_RF_INST = [
        "Age_inscription",
        "UC_inscrites_S1", "UC_validees_S1", "Note_S1",
        "UC_inscrites_S2", "UC_validees_S2", "Note_S2",
        "Frais_scolarite_a_jour",
        "Dettes_universitaires",
        "Bourse_institutionnelle",
        "Etudiant_deplace",
        "Etudiant_international",
        "Taux_chomage",
        "Taux_inflation",
        "PIB",
    ]

    # ── Valeurs autorisées par colonne (pour le template) ─
    VALEURS_AUTORISEES = {
        "1.Tranche d'âge":                   "18–20 ans | 21–23 ans | 24 ans ou plus",
        "2.Genre":                            "Femme | Homme",
        "3.Niveau d'étude actuel":            "Licence 1 (L1) | Licence 2 (L2) | Licence 3 (L3) | Master (M1/M2)",
        "4.Quelle est votre moyenne générale la plus récente ?":
                                              "Moins de 10 | Entre 10 et 12 | Entre 12 et 14 | Plus de 14",
        "5.Combien de matières avez-vous échouées lors de la dernière période académique ?":
                                              "Aucune | 1–2 | 3 ou plus",
        "6.À quelle fréquence êtes-vous absent(e) aux cours ?":
                                              "Rarement | Parfois | Souvent",
        "7.Combien d'heures étudiez-vous en moyenne par semaine (hors cours) ?":
                                              "Moins de 5 heures | Entre 5 et 10 heures | Plus de 10 heures",
        "8.Votre participation en classe est généralement":
                                              "Moyenne | Élevée",
        "9.Rendez-vous vos devoirs et travaux dans les délais ?":
                                              "Jamais | Parfois | Toujours",
        "10.À quel point procrastinez-vous dans vos études ?":
                                              "Peu | Moyennement | Beaucoup",
        "11.Je me sens motivé(e) dans mes études":
                                              "Pas du tout | Peu | Motivé(e) | Très motivé(e)",
        "12.Mon niveau de stress lié aux études est :":
                                              "Très faible | Faible | Moyen | Élevé",
        "13.J'ai confiance en ma capacité à réussir mes études":
                                              "Peu | Moyennement | Confiant(e) | Très confiant(e)",
        "14.Je ressens une fatigue mentale liée aux études":
                                              "Très faible | Faible | Moyenne | Élevée | Très élevée",
        "15.Disposez-vous d'un accès régulier à Internet pour vos études ?":
                                              "Oui | Non",
        "16.Votre environnement de travail à domicile est :":
                                              "Favorable | Moyennement favorable | Défavorable",
        "17.Le soutien de votre entourage (famille/proches) dans vos études est :":
                                              "Fort | Moyen | Faible",
        "18.Statut_Boursier":                "Oui | Non",
        "19.Activite_Salariee":              "Aucune | Moins de 15h/semaine | Plus de 15h/semaine",
        "20.Niveau_Etude_Parents":           "Bac ou moins | Licence | Master ou plus",
        "21.Type_Bac":                       "Général | Technologique | Professionnel",
        "22.Mention_Bac":                    "Sans mention | Assez bien | Bien | Très bien",
        "23.Annees_Retard":                  "0 | 1 | 2 | 3+",
        "24.Freq_Connexion_LMS":             "Quotidienne | Hebdomadaire | Rare",
        "25.Temps_Trajet_Minutes":           "Moins de 30 | 30-60 | Plus de 60",
        "26.Type_Logement":                  "Chez les parents | Logement indépendant | Cité Universitaire / Colocation",
        "Age_inscription":                   "Entier (ex: 20)",
        "UC_inscrites_S1":                   "Entier 0-12",
        "UC_validees_S1":                    "Entier 0-12",
        "Note_S1":                           "Décimal 0-20 (ex: 12.5)",
        "UC_inscrites_S2":                   "Entier 0-12",
        "UC_validees_S2":                    "Entier 0-12",
        "Note_S2":                           "Décimal 0-20 (ex: 12.5)",
        "Frais_scolarite_a_jour":            "Oui | Non",
        "Dettes_universitaires":             "Oui | Non",
        "Bourse_institutionnelle":           "Oui | Non",
        "Etudiant_deplace":                  "Oui | Non",
        "Etudiant_international":            "Oui | Non",
        "Taux_chomage":                      "Décimal (ex: 11.1)",
        "Taux_inflation":                    "Décimal (ex: 1.4)",
        "PIB":                               "Décimal (ex: 1.74)",
    }

    @st.cache_data
    def generer_template_xlsx() -> bytes:
        toutes_colonnes = ["ID_Etudiant"] + COLONNES_RF + COLONNES_XGB
        ligne_valeurs   = {col: VALEURS_AUTORISEES.get(col, "") for col in toutes_colonnes}
        ligne_valeurs["ID_Etudiant"] = "Ex: ETU-001"
        exemple_rf = {
            "1.Tranche d'âge": "18–20 ans",
            "2.Genre": "Femme",
            "3.Niveau d'étude actuel": "Licence 1 (L1)",
            "4.Quelle est votre moyenne générale la plus récente ?": "Entre 10 et 12",
            "5.Combien de matières avez-vous échouées lors de la dernière période académique ?": "1–2",
            "6.À quelle fréquence êtes-vous absent(e) aux cours ?": "Parfois",
            "7.Combien d'heures étudiez-vous en moyenne par semaine (hors cours) ?": "Entre 5 et 10 heures",
            "8.Votre participation en classe est généralement": "Moyenne",
            "9.Rendez-vous vos devoirs et travaux dans les délais ?": "Parfois",
            "10.À quel point procrastinez-vous dans vos études ?": "Moyennement",
            "11.Je me sens motivé(e) dans mes études": "Peu",
            "12.Mon niveau de stress lié aux études est :": "Élevé",
            "13.J'ai confiance en ma capacité à réussir mes études": "Moyennement",
            "14.Je ressens une fatigue mentale liée aux études": "Élevée",
            "15.Disposez-vous d'un accès régulier à Internet pour vos études ?": "Oui",
            "16.Votre environnement de travail à domicile est :": "Moyennement favorable",
            "17.Le soutien de votre entourage (famille/proches) dans vos études est :": "Moyen",
            "18.Statut_Boursier": "Oui",
            "19.Activite_Salariee": "Moins de 15h/semaine",
            "20.Niveau_Etude_Parents": "Bac ou moins",
            "21.Type_Bac": "Général",
            "22.Mention_Bac": "Assez bien",
            "23.Annees_Retard": "0",
            "24.Freq_Connexion_LMS": "Hebdomadaire",
            "25.Temps_Trajet_Minutes": "30-60",
            "26.Type_Logement": "Cité Universitaire / Colocation",
        }
        exemple_inst = {
            "Age_inscription": 19, "UC_inscrites_S1": 6, "UC_validees_S1": 4, "Note_S1": 11.0,
            "UC_inscrites_S2": 6, "UC_validees_S2": 4, "Note_S2": 10.5,
            "Frais_scolarite_a_jour": "Oui", "Dettes_universitaires": "Non",
            "Bourse_institutionnelle": "Oui", "Etudiant_deplace": "Non",
            "Etudiant_international": "Non", "Taux_chomage": 11.1,
            "Taux_inflation": 1.4, "PIB": 1.74,
        }
        ligne_exemple = {"ID_Etudiant": "ETU-001", **exemple_rf, **exemple_inst}
        df_tpl = pd.DataFrame(
            [ligne_valeurs, ligne_exemple],
            index=["[VALEURS POSSIBLES]", "[EXEMPLE]"]
        )

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_tpl.to_excel(writer, sheet_name="Template", index=True)
            wb   = writer.book
            ws   = writer.sheets["Template"]
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # ── Styles ────────────────────────────────────────
            fill_header  = PatternFill("solid", fgColor="1A3A6B")   # bleu-roi
            fill_valeurs = PatternFill("solid", fgColor="FFF8E8")   # jaune pâle
            fill_exemple = PatternFill("solid", fgColor="EDF9F4")   # vert pâle
            font_blanc   = Font(color="FFFFFF", bold=True, size=10)
            font_valeurs = Font(color="7B5C00", bold=False, size=9, italic=True)
            font_exemple = Font(color="155843", bold=True, size=10)
            font_col     = Font(color="FFFFFF", bold=True, size=9)
            border_thin  = Border(
                left=Side(style="thin", color="C0C8D8"),
                right=Side(style="thin", color="C0C8D8"),
                top=Side(style="thin", color="C0C8D8"),
                bottom=Side(style="thin", color="C0C8D8"),
            )
            align_wrap   = Alignment(wrap_text=True, vertical="top")
            align_center = Alignment(horizontal="center", vertical="center")

            max_col = ws.max_column
            max_row = ws.max_row

            # ── Ligne 1 : en-têtes de colonnes ─────────────
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill   = fill_header
                cell.font   = font_col
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.border = border_thin
            ws.row_dimensions[1].height = 40

            # ── Ligne 2 : valeurs possibles ─────────────────
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=2, column=col_idx)
                cell.fill      = fill_valeurs
                cell.font      = font_valeurs
                cell.alignment = align_wrap
                cell.border    = border_thin
            ws.row_dimensions[2].height = 55

            # ── Ligne 3 : exemple ───────────────────────────
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.fill      = fill_exemple
                cell.font      = font_exemple
                cell.alignment = align_center
                cell.border    = border_thin
            ws.row_dimensions[3].height = 22

            # ── Largeur des colonnes ─────────────────────────
            ws.column_dimensions["A"].width = 22   # index / ID
            for col_idx in range(2, max_col + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

            # ── Figer la première colonne et la première ligne ─
            ws.freeze_panes = "B2"

            # ── Listes déroulantes (DataValidation) ──────────
            DROPDOWN_CHOICES = {
                "1.Tranche d'âge":               ["18–20 ans", "21–23 ans", "24 ans ou plus"],
                "2.Genre":                       ["Femme", "Homme"],
                "3.Niveau d'étude actuel":       ["Licence 1 (L1)", "Licence 2 (L2)", "Licence 3 (L3)", "Master (M1/M2)"],
                "4.Quelle est votre moyenne générale la plus récente ?":
                                                 ["Moins de 10", "Entre 10 et 12", "Entre 12 et 14", "Plus de 14"],
                "5.Combien de matières avez-vous échouées lors de la dernière période académique ?":
                                                 ["Aucune", "1–2", "3 ou plus"],
                "6.À quelle fréquence êtes-vous absent(e) aux cours ?":
                                                 ["Rarement", "Parfois", "Souvent"],
                "7.Combien d'heures étudiez-vous en moyenne par semaine (hors cours) ?":
                                                 ["Moins de 5 heures", "Entre 5 et 10 heures", "Plus de 10 heures"],
                "8.Votre participation en classe est généralement":
                                                 ["Moyenne", "Élevée"],
                "9.Rendez-vous vos devoirs et travaux dans les délais ?":
                                                 ["Jamais", "Parfois", "Toujours"],
                "10.À quel point procrastinez-vous dans vos études ?":
                                                 ["Peu", "Moyennement", "Beaucoup"],
                "11.Je me sens motivé(e) dans mes études":
                                                 ["Pas du tout", "Peu", "Motivé(e)", "Très motivé(e)"],
                "12.Mon niveau de stress lié aux études est :":
                                                 ["Très faible", "Faible", "Moyen", "Élevé"],
                "13.J'ai confiance en ma capacité à réussir mes études":
                                                 ["Peu", "Moyennement", "Confiant(e)", "Très confiant(e)"],
                "14.Je ressens une fatigue mentale liée aux études":
                                                 ["Très faible", "Faible", "Moyenne", "Élevée", "Très élevée"],
                "15.Disposez-vous d'un accès régulier à Internet pour vos études ?":
                                                 ["Oui", "Non"],
                "16.Votre environnement de travail à domicile est :":
                                                 ["Favorable", "Moyennement favorable", "Défavorable"],
                "17.Le soutien de votre entourage (famille/proches) dans vos études est :":
                                                 ["Fort", "Moyen", "Faible"],
                "18.Statut_Boursier":            ["Oui", "Non"],
                "19.Activite_Salariee":          ["Aucune", "Moins de 15h/semaine", "Plus de 15h/semaine"],
                "20.Niveau_Etude_Parents":       ["Bac ou moins", "Licence", "Master ou plus"],
                "21.Type_Bac":                   ["Général", "Technologique", "Professionnel"],
                "22.Mention_Bac":                ["Sans mention", "Assez bien", "Bien", "Très bien"],
                "23.Annees_Retard":              ["0", "1", "2", "3+"],
                "24.Freq_Connexion_LMS":         ["Quotidienne", "Hebdomadaire", "Rare"],
                "25.Temps_Trajet_Minutes":       ["Moins de 30", "30-60", "Plus de 60"],
                "26.Type_Logement":              ["Chez les parents", "Logement indépendant", "Cité Universitaire / Colocation"],
                "Frais_scolarite_a_jour":        ["Oui", "Non"],
                "Dettes_universitaires":         ["Oui", "Non"],
                "Bourse_institutionnelle":       ["Oui", "Non"],
                "Etudiant_deplace":              ["Oui", "Non"],
                "Etudiant_international":        ["Oui", "Non"],
            }

            from openpyxl.worksheet.datavalidation import DataValidation

            # Construire le mapping nom_colonne → lettre Excel via la ligne d'en-tête
            col_letter_map = {}
            for ci in range(1, max_col + 1):
                header_val = ws.cell(row=1, column=ci).value
                if header_val:
                    col_letter_map[str(header_val)] = get_column_letter(ci)

            # Les lignes 1-3 sont réservées (en-tête + template) ;
            # les données utilisateur commencent à la ligne 4
            DATA_START = 4
            DATA_END   = 1000

            for col_name, choices in DROPDOWN_CHOICES.items():
                col_ltr = col_letter_map.get(col_name)
                if not col_ltr:
                    continue
                choices_formula = '"' + ",".join(choices) + '"'
                dv = DataValidation(
                    type="list",
                    formula1=choices_formula,
                    allow_blank=True,
                    showDropDown=False,        # False = flèche déroulante visible
                    showErrorMessage=True,
                    errorTitle="Valeur non reconnue",
                    error="Utilisez la liste déroulante pour sélectionner une valeur valide.",
                    showInputMessage=True,
                    promptTitle=col_name[:32],
                    prompt="↓ Cliquez sur la flèche pour choisir une option.",
                )
                dv.sqref = f"{col_ltr}{DATA_START}:{col_ltr}{DATA_END}"
                ws.add_data_validation(dv)

        return buf.getvalue()

    # ── Zone d'upload ──────────────────────────────────────
    st.markdown("""
    <div class="upload-zone">
        <div class="upload-title">📂 Import d'un fichier d'étudiants</div>
        <div class="upload-hint">Format accepté : CSV (.csv) ou Excel (.xlsx) — encodage UTF-8 recommandé</div>
    </div>
    """, unsafe_allow_html=True)

    col_up, col_tpl = st.columns([2, 1])
    with col_up:
        fichier = st.file_uploader(
            "Choisir un fichier",
            type=["csv", "xlsx"],
            label_visibility="collapsed",
        )
    with col_tpl:
        st.download_button(
            label="⬇️  Télécharger le template Excel",
            data=generer_template_xlsx(),
            file_name="template_alerte_precoce.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Téléchargez ce modèle Excel, remplissez-le avec vos données, puis importez-le ici.",
        )

    if fichier is not None:
        # ── Lecture du fichier ─────────────────────────────
        try:
            if fichier.name.endswith(".xlsx"):
                df_upload = pd.read_excel(fichier)
            else:
                df_upload = pd.read_csv(fichier, encoding="utf-8-sig")
        except Exception as e:
            st.error(f"❌ Impossible de lire le fichier : {e}")
            st.stop()

        # ── Suppression des lignes de métadonnées du template ──
        # Le template écrit l'index pandas en colonne A (sans nom ou nommée "Unnamed: 0").
        # Ces lignes ont pour index "[VALEURS POSSIBLES]" ou "[EXEMPLE]".
        # On filtre sur toutes les colonnes de type objet pour détecter ces marqueurs.
        MARQUEURS_TEMPLATE = {"[valeurs possibles]", "[exemple]", "ex: etu-001"}
        col_index_unnamed = [c for c in df_upload.columns if str(c).lower().startswith("unnamed")]
        if col_index_unnamed:
            # Colonne index exportée par pandas (colonne A du xlsx)
            col_idx_name = col_index_unnamed[0]
            masque = df_upload[col_idx_name].astype(str).str.strip().str.lower().isin(MARQUEURS_TEMPLATE)
            df_upload = df_upload[~masque].drop(columns=[col_idx_name]).reset_index(drop=True)
        else:
            # Fallback : filtrer via toutes les colonnes string
            def _est_ligne_template(row):
                return any(str(v).strip().lower() in MARQUEURS_TEMPLATE for v in row.values)
            masque = df_upload.apply(_est_ligne_template, axis=1)
            df_upload = df_upload[~masque].reset_index(drop=True)

        st.success(f"✅ Fichier chargé — **{len(df_upload)} étudiant(s)** détecté(s)")

        # ── Aperçu du fichier ──────────────────────────────
        with st.expander("👁️ Aperçu des données chargées (5 premières lignes)", expanded=False):
            st.dataframe(df_upload.head(5), use_container_width=True, hide_index=True)

        # ── Validation des valeurs catégorielles ───────────
        VALEURS_VALIDES = {
            "1.Tranche d'âge":               ['18–20 ans', '21–23 ans', '24 ans ou plus'],
            "2.Genre":                       ['Femme', 'Homme'],
            "3.Niveau d'étude actuel":       ['Licence 1 (L1)', 'Licence 2 (L2)', 'Licence 3 (L3)', 'Master (M1/M2)'],
            "4.Quelle est votre moyenne générale la plus récente ?":
                                             ['Moins de 10', 'Entre 10 et 12', 'Entre 12 et 14', 'Plus de 14'],
            "5.Combien de matières avez-vous échouées lors de la dernière période académique ?":
                                             ['Aucune', '1–2', '3 ou plus'],
            "6.À quelle fréquence êtes-vous absent(e) aux cours ?":
                                             ['Rarement', 'Parfois', 'Souvent'],
            "7.Combien d'heures étudiez-vous en moyenne par semaine (hors cours) ?":
                                             ['Moins de 5 heures', 'Entre 5 et 10 heures', 'Plus de 10 heures'],
            "8.Votre participation en classe est généralement":
                                             ['Moyenne', 'Élevée'],
            "9.Rendez-vous vos devoirs et travaux dans les délais ?":
                                             ['Jamais', 'Parfois', 'Toujours'],
            "10.À quel point procrastinez-vous dans vos études ?":
                                             ['Peu', 'Moyennement', 'Beaucoup'],
            "11.Je me sens motivé(e) dans mes études":
                                             ['Pas du tout', 'Peu', 'Motivé(e)', 'Très motivé(e)'],
            "12.Mon niveau de stress lié aux études est :":
                                             ['Très faible', 'Faible', 'Moyen', 'Élevé'],
            "13.J'ai confiance en ma capacité à réussir mes études":
                                             ['Peu', 'Moyennement', 'Confiant(e)', 'Très confiant(e)'],
            "14.Je ressens une fatigue mentale liée aux études":
                                             ['Très faible', 'Faible', 'Moyenne', 'Élevée', 'Très élevée'],
            "18.Statut_Boursier":            ['Oui', 'Non'],
            "19.Activite_Salariee":          ['Aucune', 'Moins de 15h/semaine', 'Plus de 15h/semaine'],
            "20.Niveau_Etude_Parents":       ['Bac ou moins', 'Licence', 'Master ou plus'],
            "21.Type_Bac":                   ['Général', 'Technologique', 'Professionnel'],
            "22.Mention_Bac":                ['Sans mention', 'Assez bien', 'Bien', 'Très bien'],
            "23.Annees_Retard":              ['0', '1', '2', '3+'],
            "24.Freq_Connexion_LMS":         ['Quotidienne', 'Hebdomadaire', 'Rare'],
            "25.Temps_Trajet_Minutes":       ['Moins de 30', '30-60', 'Plus de 60'],
        }
        erreurs_validation = []
        for col, valides in VALEURS_VALIDES.items():
            if col not in df_upload.columns:
                continue
            invalides = df_upload[~df_upload[col].astype(str).isin(valides)]
            for row_i, row_v in invalides.iterrows():
                id_e = row_v.get("ID_Etudiant", f"ligne {row_i+1}")
                erreurs_validation.append(
                    f"**{id_e}** · colonne `{col}` · valeur reçue : `{row_v[col]}`"
                )

        if erreurs_validation:
            with st.expander(f"⚠️ {len(erreurs_validation)} valeur(s) non reconnue(s) — cliquez pour voir le détail", expanded=True):
                st.caption("Ces cellules seront imputées par la médiane du modèle. Corrigez le fichier pour de meilleurs résultats.")
                for e in erreurs_validation[:20]:
                    st.markdown(f"- {e}")
                if len(erreurs_validation) > 20:
                    st.caption(f"… et {len(erreurs_validation)-20} autres erreurs.")
        else:
            st.success("✅ Toutes les valeurs catégorielles sont valides.")

        # ── Vérification des colonnes RF obligatoires ──────
        cols_manquantes = [c for c in COLONNES_RF if c not in df_upload.columns]
        if cols_manquantes:
            st.error(f"❌ Colonnes RF manquantes dans le fichier ({len(cols_manquantes)}) :")
            st.code("\n".join(cols_manquantes))
            st.info("💡 Téléchargez le template CSV ci-dessus pour voir le format attendu.")
            st.stop()

        # ── Traitement batch ───────────────────────────────
        with st.spinner(f"Analyse de {len(df_upload)} étudiant(s) en cours..."):

            resultats = []
            explainer_batch = shap.TreeExplainer(rf_model)

            for idx, row in df_upload.iterrows():
                id_etu = str(row.get("ID_Etudiant", f"Étudiant {idx+1}"))

                # — Données RF
                donnees_rf = {col: row.get(col, "") for col in COLONNES_RF}

                try:
                    input_rf_scaled, _ = pretraiter(donnees_rf, scaler_rf, rf_features)
                    p_rf = float(rf_model.predict_proba(input_rf_scaled)[0, 1])
                except Exception:
                    p_rf = float("nan")

                # — Données RF institutionnel
                # Initialiser toutes les features à 0, puis écraser les 13 champs saisis
                row_inst   = {col: 0 for col in inst_features}
                input_inst_df = pd.DataFrame([row_inst], columns=inst_features)

                inst_map_batch = {
                    "Age at enrollment":                    "Age_inscription",
                    "Curricular units 1st sem (enrolled)":  "UC_inscrites_S1",
                    "Curricular units 1st sem (approved)":  "UC_validees_S1",
                    "Curricular units 1st sem (grade)":     "Note_S1",
                    "Curricular units 2nd sem (enrolled)":  "UC_inscrites_S2",
                    "Curricular units 2nd sem (approved)":  "UC_validees_S2",
                    "Curricular units 2nd sem (grade)":     "Note_S2",
                    "Tuition fees up to date_1":            "Frais_scolarite_a_jour",
                    "Debtor_1":                             "Dettes_universitaires",
                    "Scholarship holder_1":                 "Bourse_institutionnelle",
                    "Displaced_1":                          "Etudiant_deplace",
                    "International_1":                      "Etudiant_international",
                    "Unemployment rate":                    "Taux_chomage",
                    "Inflation rate":                       "Taux_inflation",
                    "GDP":                                  "PIB",
                }

                for col_inst, col_csv in inst_map_batch.items():
                    if col_inst in input_inst_df.columns and col_csv in df_upload.columns:
                        val = row.get(col_csv, None)
                        if pd.notna(val):
                            if str(val).strip().lower() in ("oui", "yes", "1"):
                                val = 1
                            elif str(val).strip().lower() in ("non", "no", "0"):
                                val = 0
                            try:
                                input_inst_df[col_inst] = float(val)
                            except (ValueError, TypeError):
                                pass

                try:
                    p_inst = float(rf_inst.predict_proba(input_inst_df)[0, 1])
                except Exception:
                    p_inst = float("nan")

                # — Score final fusionné
                w_rf, w_inst = 0.60, 0.40
                if pd.isna(p_rf) and pd.isna(p_inst):
                    risque = float("nan")
                elif pd.isna(p_rf):
                    risque = p_inst
                elif pd.isna(p_inst):
                    risque = p_rf
                else:
                    risque = p_rf * w_rf + p_inst * w_inst

                statut = "🔴 ALERTE" if (not pd.isna(risque) and risque >= seuil_fusion) else "🟢 OK"

                # — Recommandations principales (SHAP)
                try:
                    shap_raw_b    = explainer_batch.shap_values(input_rf_scaled)
                    sv_b          = extraire_sv(shap_raw_b)
                    feat_labels_b = [traduire_feature(f) for f in rf_features]
                    recos_b       = generer_recommandations(sv_b, feat_labels_b)
                    reco_txt      = " | ".join([r[:60] + "…" for r in recos_b[:2]]) if recos_b else "—"
                except Exception:
                    reco_txt      = "—"

                resultats.append({
                    "ID Étudiant":            id_etu,
                    "Score RF (60%)":         f"{p_rf:.1%}"   if not pd.isna(p_rf)   else "Erreur",
                    "Score RF Inst. (40%)":   f"{p_inst:.1%}" if not pd.isna(p_inst) else "Erreur",
                    "Risque Final":           f"{risque:.1%}" if not pd.isna(risque)  else "Erreur",
                    "Statut":                 statut,
                    "Recommandations":        reco_txt,
                    # Valeurs brutes pour le CSV d'export
                    "_p_rf_raw":              round(p_rf,   4) if not pd.isna(p_rf)   else None,
                    "_p_inst_raw":            round(p_inst, 4) if not pd.isna(p_inst) else None,
                    "_risque_raw":            round(risque, 4) if not pd.isna(risque)  else None,
                })

        # ── Conversion en DataFrame + tri par risque décroissant ──
        df_res = pd.DataFrame(resultats)
        df_res = df_res.sort_values("_risque_raw", ascending=False, na_position="last").reset_index(drop=True)

        # ── Métriques de synthèse ──────────────────────────   
        st.markdown("---")
        st.markdown("## 📊 Résultats de l'analyse batch")

        nb_alerte   = (df_res["Statut"] == "🔴 ALERTE").sum()
        nb_ok       = (df_res["Statut"] == "🟢 OK").sum()
        nb_total    = len(df_res)
        taux_alerte = nb_alerte / nb_total * 100 if nb_total > 0 else 0

        m1, m2, m3, m4 = st.columns(4)
        with m1:
            st.markdown(f'<div class="metric-mini"><div class="metric-mini-val">{nb_total}</div><div class="metric-mini-label">Étudiants analysés</div></div>', unsafe_allow_html=True)
        with m2:
            st.markdown(f'<div class="metric-mini"><div class="metric-mini-val" style="color:#c0392b;">{nb_alerte}</div><div class="metric-mini-label">En alerte</div></div>', unsafe_allow_html=True)
        with m3:
            st.markdown(f'<div class="metric-mini"><div class="metric-mini-val" style="color:#1e7e5a;">{nb_ok}</div><div class="metric-mini-label">Profil OK</div></div>', unsafe_allow_html=True)
        with m4:
            couleur_taux = "#c0392b" if taux_alerte > 40 else "#f39c12" if taux_alerte > 20 else "#1e7e5a"
            st.markdown(f'<div class="metric-mini"><div class="metric-mini-val" style="color:{couleur_taux};">{taux_alerte:.0f}%</div><div class="metric-mini-label">Taux d\'alerte</div></div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Tableau stylisé avec coloration par statut ────
        cols_affichage = ["ID Étudiant", "Score RF (60%)", "Score RF Inst. (40%)", "Risque Final", "Statut", "Recommandations"]

        def _colorier_ligne(row):
            if row["Statut"] == "🔴 ALERTE":
                return ["background-color:#fdf0ee; color:#7b1c10"] * len(row)
            else:
                return ["background-color:#edfaf5; color:#155843"] * len(row)

        df_affich  = df_res[cols_affichage].copy()
        df_styled  = df_affich.style.apply(_colorier_ligne, axis=1)
        st.dataframe(
            df_styled,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Statut":          st.column_config.TextColumn("Statut", width="small"),
                "Recommandations": st.column_config.TextColumn("Recommandations principales", width="large"),
            }
        )

        # ── Sauvegarde dans l'historique ──────────────────
        st.session_state.historique.append({
            "type":      "batch",
            "horodatage": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
            "fichier":   fichier.name,
            "nb_total":  nb_total,
            "nb_alerte": int(nb_alerte),
            "df_res":    df_res[cols_affichage].copy(),
        })

        # ── Export CSV ─────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)

        def generer_csv_export(df: pd.DataFrame) -> bytes:
            cols_export = ["ID Étudiant", "_p_rf_raw", "_p_inst_raw", "_risque_raw", "Statut", "Recommandations"]
            df_exp = df[cols_export].copy()
            df_exp.columns = ["ID Étudiant", "Score RF", "Score RF Inst.", "Risque Final", "Statut", "Recommandations"]
            buf = io.BytesIO()
            df_exp.to_csv(buf, index=False, encoding="utf-8-sig")
            return buf.getvalue()

        col_exp1, col_exp2, col_exp3 = st.columns([1, 2, 1])
        with col_exp2:
            st.download_button(
                label="⬇️  Exporter les résultats (CSV)",
                data=generer_csv_export(df_res),
                file_name="resultats_alerte_precoce.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # Arrêt de l'exécution ici — ne pas afficher le formulaire manuel
    st.stop()

# ═════════════════════════════════════════════════════════════
# ── MODE SAISIE MANUELLE ──────────────────────────────────────
# ═════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────
# FORMULAIRE EN ONGLETS
# ─────────────────────────────────────────────────────────────
st.markdown("### 📋 Saisie du profil étudiant")
st.caption("Renseignez les informations de l'étudiant dans les 5 onglets ci-dessous.")

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "👤 Profil général",
    "📚 Parcours académique",
    "🧠 Profil psychologique",
    "🏠 Contexte de vie",
    "🏛️ Données institutionnelles",
])

# ── Onglet 1 : Profil général ──────────────────────────────
with tab1:
    st.markdown('<div class="section-card"><div class="section-title">👤 Informations générales</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        age         = st.selectbox("Tranche d'âge",        ['18–20 ans', '21–23 ans', '24 ans ou plus'],        key="ss_age")
        genre       = st.selectbox("Genre",                 ['Femme', 'Homme'],                                  key="ss_genre")
    with c2:
        niveau      = st.selectbox("Niveau d'étude actuel", ['Licence 1 (L1)', 'Licence 2 (L2)', 'Licence 3 (L3)', 'Master (M1/M2)'], key="ss_niveau")
        type_bac    = st.selectbox("Type de Bac",           ['Général', 'Technologique', 'Professionnel'],       key="ss_type_bac")
    mention_bac     = st.selectbox("Mention au Bac",        ['Sans mention', 'Assez bien', 'Bien', 'Très bien'], key="ss_mention_bac")
    st.markdown('</div>', unsafe_allow_html=True)

# ── Onglet 2 : Parcours académique ─────────────────────────
with tab2:
    st.markdown('<div class="section-card"><div class="section-title">📚 Résultats et comportements académiques</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        moyenne       = st.selectbox("Moyenne générale",       ['Moins de 10', 'Entre 10 et 12', 'Entre 12 et 14', 'Plus de 14'], key="ss_moyenne")
        matieres      = st.selectbox("Matières échouées",      ['Aucune', '1–2', '3 ou plus'],                                    key="ss_matieres")
        absences      = st.selectbox("Fréquence des absences", ['Rarement', 'Parfois', 'Souvent'],                                key="ss_absences")
        annees_ret    = st.selectbox("Années de retard",       ['0', '1', '2', '3+'],                                            key="ss_annees_ret")
    with c2:
        heures        = st.selectbox("Heures d'étude / semaine",  ['Moins de 5 heures', 'Entre 5 et 10 heures', 'Plus de 10 heures'], key="ss_heures")
        participation = st.selectbox("Participation en classe",   ['Moyenne', 'Élevée'],                                              key="ss_participation")
        devoirs       = st.selectbox("Devoirs dans les délais",   ['Jamais', 'Parfois', 'Toujours'],                                  key="ss_devoirs")
    st.markdown('</div>', unsafe_allow_html=True)

# ── Onglet 3 : Profil psychologique ────────────────────────
with tab3:
    st.markdown('<div class="section-card"><div class="section-title">🧠 État psychologique et engagement</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        procrastin  = st.selectbox("Niveau de procrastination",  ['Peu', 'Moyennement', 'Beaucoup'],                          key="ss_procrastin")
        motivation  = st.selectbox("Motivation dans les études", ['Très motivé(e)', 'Motivé(e)', 'Peu', 'Pas du tout'],       key="ss_motivation")
        stress      = st.selectbox("Niveau de stress",           ['Très faible', 'Faible', 'Moyen', 'Élevé'],                 key="ss_stress")
    with c2:
        confiance   = st.selectbox("Confiance en soi",  ['Très confiant(e)', 'Confiant(e)', 'Moyennement', 'Peu'],            key="ss_confiance")
        fatigue     = st.selectbox("Fatigue mentale",   ['Très faible', 'Faible', 'Moyenne', 'Élevée', 'Très élevée'],       key="ss_fatigue")
    st.markdown('</div>', unsafe_allow_html=True)

# ── Onglet 4 : Contexte de vie ──────────────────────────────
with tab4:
    st.markdown('<div class="section-card"><div class="section-title">🏠 Contexte socio-économique et environnement</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        statut_bours  = st.selectbox("Statut boursier",            ['Non', 'Oui'],                                                          key="ss_statut_bours")
        activite_sal  = st.selectbox("Activité salariée",          ['Aucune', 'Moins de 15h/semaine', 'Plus de 15h/semaine'],                key="ss_activite_sal")
        niv_parents   = st.selectbox("Niveau d'étude des parents", ['Bac ou moins', 'Licence', 'Master ou plus'],                           key="ss_niv_parents")
        internet      = st.selectbox("Accès Internet régulier",    ['Oui', 'Non'],                                                          key="ss_internet")
    with c2:
        environnement = st.selectbox("Environnement de travail",   ['Favorable', 'Moyennement favorable', 'Défavorable'],                   key="ss_environnement")
        soutien       = st.selectbox("Soutien de l'entourage",     ['Fort', 'Moyen', 'Faible'],                                             key="ss_soutien")
        freq_lms      = st.selectbox("Fréquence connexion LMS",    ['Quotidienne', 'Hebdomadaire', 'Rare'],                                 key="ss_freq_lms")
        trajet        = st.selectbox("Temps de trajet (min)",      ['Moins de 30', '30-60', 'Plus de 60'],                                  key="ss_trajet")
        logement      = st.selectbox("Type de logement",           ['Chez les parents', 'Logement indépendant', 'Cité Universitaire / Colocation'], key="ss_logement")
    st.markdown('</div>', unsafe_allow_html=True)

# ── Onglet 5 : Données institutionnelles (13 champs) ────────
with tab5:
    st.markdown('<div class="section-card"><div class="section-title">🏛️ Données institutionnelles — Modèle RF institutionnel</div>', unsafe_allow_html=True)
    st.caption(
        "Ces 13 variables alimentent le modèle RF institutionnel (40 % du score final). "
        "Les autres variables du dataset UCI sont initialisées à 0 (valeur neutre pour les features OHE)."
    )

    # ── Bloc 1 : Résultats académiques officiels ──────────
    st.markdown("**📊 Résultats académiques officiels**")
    ci1, ci2, ci3 = st.columns(3)
    with ci1:
        age_enrol   = st.number_input("Âge à l'inscription",
                                      min_value=17, max_value=60, value=20,
                                      help="Âge de l'étudiant au moment de son inscription initiale.",
                                      key="ss_age_enrol")
        curr_1_enr  = st.number_input("UC inscrites S1",
                                      min_value=0, max_value=12, value=6,
                                      help="Nombre d'unités curriculaires dans lesquelles l'étudiant s'est inscrit au 1er semestre.",
                                      key="ss_curr_1_enr")
        curr_1_app  = st.number_input("UC validées S1",
                                      min_value=0, max_value=12, value=5,
                                      help="Nombre d'unités curriculaires réussies au 1er semestre.",
                                      key="ss_curr_1_app")
    with ci2:
        curr_1_note = st.number_input("Note moyenne S1",
                                      min_value=0.0, max_value=20.0, value=12.0, step=0.5,
                                      help="Moyenne des notes obtenues au 1er semestre (sur 20).",
                                      key="ss_curr_1_note")
        curr_2_enr  = st.number_input("UC inscrites S2",
                                      min_value=0, max_value=12, value=6,
                                      help="Nombre d'unités curriculaires dans lesquelles l'étudiant s'est inscrit au 2e semestre.",
                                      key="ss_curr_2_enr")
        curr_2_app  = st.number_input("UC validées S2",
                                      min_value=0, max_value=12, value=5,
                                      help="Nombre d'unités curriculaires réussies au 2e semestre.",
                                      key="ss_curr_2_app")
    with ci3:
        curr_2_note = st.number_input("Note moyenne S2",
                                      min_value=0.0, max_value=20.0, value=12.0, step=0.5,
                                      help="Moyenne des notes obtenues au 2e semestre (sur 20).",
                                      key="ss_curr_2_note")

    st.markdown("---")

    # ── Bloc 2 : Situation administrative ─────────────────
    st.markdown("**📋 Situation administrative**")
    ca1, ca2, ca3 = st.columns(3)
    with ca1:
        frais_ok    = st.selectbox("Frais de scolarité à jour",
                                   ['Oui', 'Non'],
                                   help="L'étudiant a-t-il réglé ses frais de scolarité pour l'année en cours ?",
                                   key="ss_frais_ok")
        dettes      = st.selectbox("Dettes universitaires",
                                   ['Non', 'Oui'],
                                   help="L'étudiant a-t-il des dettes auprès de l'établissement ?",
                                   key="ss_dettes")
    with ca2:
        bourse_inst = st.selectbox("Bénéficiaire d'une bourse",
                                   ['Non', 'Oui'],
                                   help="L'étudiant bénéficie-t-il d'une bourse d'études institutionnelle ?",
                                   key="ss_bourse_inst")
        deplace     = st.selectbox("Étudiant déplacé",
                                   ['Non', 'Oui'],
                                   help="L'étudiant a-t-il dû quitter son lieu de résidence pour étudier ?",
                                   key="ss_deplace")
    with ca3:
        international = st.selectbox("Étudiant international",
                                     ['Non', 'Oui'],
                                     help="L'étudiant est-il de nationalité étrangère ?",
                                     key="ss_international")

    st.markdown("---")

    # ── Bloc 3 : Indicateurs macroéconomiques ─────────────
    st.markdown("**🌍 Indicateurs macroéconomiques** *(pré-remplis avec les valeurs courantes)*")
    cm1, cm2, cm3 = st.columns(3)
    with cm1:
        chomage     = st.number_input("Taux de chômage national (%)",
                                      min_value=0.0, max_value=30.0, value=11.1, step=0.1,
                                      help="Taux de chômage national au moment de l'inscription (valeur UCI médiane : 11.1%).",
                                      key="ss_chomage")
    with cm2:
        inflation   = st.number_input("Taux d'inflation (%)",
                                      min_value=-5.0, max_value=20.0, value=1.4, step=0.1,
                                      help="Taux d'inflation national au moment de l'inscription (valeur UCI médiane : 1.4%).",
                                      key="ss_inflation")
    with cm3:
        gdp         = st.number_input("PIB (taux de croissance, %)",
                                      min_value=-10.0, max_value=10.0, value=1.74, step=0.01,
                                      help="Taux de croissance du PIB national (valeur UCI médiane : 1.74%).",
                                      key="ss_gdp")

    st.markdown(
        '<div class="info-mediane">ℹ️ Les variables non saisies ici (mode de candidature, filière, '
        'qualifications parentales, etc.) sont initialisées à 0 — valeur neutre pour les features '
        'binaires issues du One-Hot Encoding du dataset UCI.</div>',
        unsafe_allow_html=True
    )
    st.markdown('</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# BOUTON ANALYSER
# ─────────────────────────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
col_btn = st.columns([1, 2, 1])[1]
with col_btn:
    analyser = st.button("🔍 Analyser le risque de décrochage")

# ─────────────────────────────────────────────────────────────
# RÉSULTATS
# ─────────────────────────────────────────────────────────────
if analyser:
    with st.spinner("Analyse en cours..."):

        # ── 1. Dictionnaire étudiant RF ───────────────────────
        donnees_etudiant = {
            "1.Tranche d'âge"                                                                    : age,
            '2.Genre'                                                                            : genre,
            "3.Niveau d'étude actuel"                                                            : niveau,
            '4.Quelle est votre moyenne générale la plus récente ?'                              : moyenne,
            '5.Combien de matières avez-vous échouées lors de la dernière période académique ?'  : matieres,
            '6.À quelle fréquence êtes-vous absent(e) aux cours ?'                              : absences,
            "7.Combien d'heures étudiez-vous en moyenne par semaine (hors cours) ?"             : heures,
            '8.Votre participation en classe est généralement'                                  : participation,
            '9.Rendez-vous vos devoirs et travaux dans les délais ?'                            : devoirs,
            '10.À quel point procrastinez-vous dans vos études ?'                               : procrastin,
            '11.Je me sens motivé(e) dans mes études'                                           : motivation,
            '12.Mon niveau de stress lié aux études est :'                                      : stress,
            "13.J'ai confiance en ma capacité à réussir mes études"                             : confiance,
            '14.Je ressens une fatigue mentale liée aux études'                                 : fatigue,
            "15.Disposez-vous d'un accès régulier à Internet pour vos études ?"                 : internet,
            '16.Votre environnement de travail à domicile est :'                                : environnement,
            '17.Le soutien de votre entourage (famille/proches) dans vos études est :'          : soutien,
            '18.Statut_Boursier'                                                                : statut_bours,
            '19.Activite_Salariee'                                                              : activite_sal,
            '20.Niveau_Etude_Parents'                                                           : niv_parents,
            '21.Type_Bac'                                                                       : type_bac,
            '22.Mention_Bac'                                                                    : mention_bac,
            '23.Annees_Retard'                                                                  : annees_ret,
            '24.Freq_Connexion_LMS'                                                             : freq_lms,
            '25.Temps_Trajet_Minutes'                                                           : trajet,
            '26.Type_Logement'                                                                  : logement,
        }

        # ── 2. Prétraitement RF ───────────────────────────────
        input_rf_scaled, _ = pretraiter(donnees_etudiant, scaler_rf, rf_features)

        # ── 3. Input RF institutionnel — 13 champs saisis ────
        row_inst      = {col: 0 for col in inst_features}
        input_inst_df = pd.DataFrame([row_inst], columns=inst_features)

        overrides = [
            ('Age at enrollment',                        age_enrol),
            ('Curricular units 1st sem (enrolled)',       curr_1_enr),
            ('Curricular units 1st sem (approved)',       curr_1_app),
            ('Curricular units 1st sem (grade)',          curr_1_note),
            ('Curricular units 2nd sem (enrolled)',       curr_2_enr),
            ('Curricular units 2nd sem (approved)',       curr_2_app),
            ('Curricular units 2nd sem (grade)',          curr_2_note),
            ('Tuition fees up to date_1',                 1 if frais_ok    == 'Oui' else 0),
            ('Debtor_1',                                  1 if dettes      == 'Oui' else 0),
            ('Scholarship holder_1',                      1 if bourse_inst == 'Oui' else 0),
            ('Displaced_1',                               1 if deplace     == 'Oui' else 0),
            ('International_1',                           1 if international == 'Oui' else 0),
            ('Unemployment rate',                         chomage),
            ('Inflation rate',                            inflation),
            ('GDP',                                       gdp),
        ]
        for col_name, val in overrides:
            if col_name in input_inst_df.columns:
                input_inst_df[col_name] = val

        # ── 4. Prédictions ─────────────────────────────────────
        p_rf   = float(rf_model.predict_proba(input_rf_scaled)[0, 1])
        p_inst = float(rf_inst.predict_proba(pretraiter_inst(input_inst_df, inst_features))[0, 1])

        w_rf, w_inst = 0.60, 0.40
        risque_final = p_rf * w_rf + p_inst * w_inst
        est_alerte   = risque_final >= seuil_fusion
        pct          = int(risque_final * 100)

        # ── 5. SHAP comportemental ─────────────────────────────
        explainer      = shap.TreeExplainer(rf_model)
        shap_raw       = explainer.shap_values(input_rf_scaled)
        sv             = extraire_sv(shap_raw)
        feat_labels    = [traduire_feature(f) for f in rf_features]
        n              = min(len(sv), len(feat_labels))
        sv             = sv[:n]
        feat_labels    = feat_labels[:n]

        # ── 6. SHAP institutionnel ─────────────────────────────
        input_inst_aligned = pretraiter_inst(input_inst_df, inst_features)
        explainer_inst     = shap.TreeExplainer(rf_inst)
        shap_raw_inst      = explainer_inst.shap_values(input_inst_aligned)
        sv_inst            = extraire_sv(shap_raw_inst)
        feat_labels_inst   = [traduire_feature(f) for f in inst_features]
        n_inst             = min(len(sv_inst), len(feat_labels_inst))
        sv_inst            = sv_inst[:n_inst]
        feat_labels_inst   = feat_labels_inst[:n_inst]

    # ─────────────────────────────────────────────────────────
    # AFFICHAGE RÉSULTATS
    # ─────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("## 📊 Résultats de l'analyse")

    card_class  = "result-alerte"    if est_alerte else "result-securite"
    statut_txt  = "🔴 ALERTE — Risque élevé de décrochage" if est_alerte else "🟢 SÉCURITÉ — Risque faible de décrochage"
    jauge_class = "jauge-fill-rouge" if est_alerte else "jauge-fill-vert"

    st.markdown(f"""
    <div class="{card_class}">
        <div class="result-title">{statut_txt}</div>
        <div class="result-score">Score de risque global : <strong>{pct}%</strong></div>
        <div class="jauge-container">
            <div class="{jauge_class}" style="width:{pct}%"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    col_pond, col_shap, col_reco = st.columns([1, 1.6, 1])

    # ── Pondération ────────────────────────────────────────
    with col_pond:
        st.markdown("#### ⚖️ Détail de la pondération")
        st.markdown(f"""
        <div class="pond-card">
            <div>
                <div class="pond-label">🌿 Score comportemental (RF)</div>
                <div class="pond-detail">{p_rf:.1%} × 60%</div>
            </div>
            <div class="pond-value">{p_rf*w_rf:.1%}</div>
        </div>
        <div class="pond-card">
            <div>
                <div class="pond-label">🏛️ Score institutionnel (RF)</div>
                <div class="pond-detail">{p_inst:.1%} × 40%</div>
            </div>
            <div class="pond-value">{p_inst*w_inst:.1%}</div>
        </div>
        <div class="pond-card" style="background:#e8ecf0;border:1px solid #c5cdd8;">
            <div>
                <div class="pond-label"><strong>Score final</strong></div>
                <div class="pond-detail">RF comport. + RF inst. · Seuil : {seuil_fusion:.0%}</div>
            </div>
            <div class="pond-value" style="font-size:1.2rem;">{risque_final:.1%}</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        m1, m2 = st.columns(2)
        with m1:
            st.markdown(f'<div class="metric-mini"><div class="metric-mini-val">{p_rf:.0%}</div><div class="metric-mini-label">Score RF comport.</div></div>', unsafe_allow_html=True)
        with m2:
            st.markdown(f'<div class="metric-mini"><div class="metric-mini-val">{p_inst:.0%}</div><div class="metric-mini-label">Score RF inst.</div></div>', unsafe_allow_html=True)

    # ── Graphique SHAP — deux graphiques côte à côte ──────
    with col_shap:
        st.markdown("#### 🔍 Explication du score (SHAP)")
        st.caption("Variables ayant le plus influencé chaque prédiction.")

        def _plot_shap(sv_arr, labels, title):
            top_n        = 8
            indices      = [int(i) for i in np.argsort(np.abs(sv_arr))[::-1][:top_n]]
            top_sv       = np.array([sv_arr[i] for i in indices], dtype=float)
            top_lbl      = [labels[i]           for i in indices]
            sorted_order = list(np.argsort(top_sv))
            sorted_sv    = np.array([top_sv[i]  for i in sorted_order], dtype=float)
            sorted_lbl   = [top_lbl[i]           for i in sorted_order]
            colors       = ['#c0392b' if v > 0 else '#1e7e5a' for v in sorted_sv]
            fig, ax = plt.subplots(figsize=(5.5, 3.8))
            fig.patch.set_facecolor('#fafbfc')
            ax.set_facecolor('#fafbfc')
            ax.barh(range(len(sorted_sv)), sorted_sv, color=colors, height=0.6, edgecolor='none')
            ax.set_yticks(range(len(sorted_lbl)))
            ax.set_yticklabels(sorted_lbl, fontsize=8, fontfamily='sans-serif')
            ax.axvline(x=0, color='#999', linewidth=0.8)
            ax.set_title(title, fontsize=9, fontweight='bold', color='#1a3a6b', pad=6)
            ax.set_xlabel("Impact sur le score de risque", fontsize=8, color='#666')
            ax.tick_params(colors='#666', labelsize=8)
            for spine in ax.spines.values():
                spine.set_visible(False)
            ax.grid(axis='x', alpha=0.3, linewidth=0.5)
            ax.legend(
                handles=[mpatches.Patch(color='#c0392b', label='↑ Augmente le risque'),
                         mpatches.Patch(color='#1e7e5a', label='↓ Réduit le risque')],
                fontsize=7, loc='lower right', framealpha=0.8
            )
            plt.tight_layout()
            return fig

        shap_c1, shap_c2 = st.columns(2)
        with shap_c1:
            st.pyplot(_plot_shap(sv, feat_labels, "🌿 RF Comportemental"), use_container_width=True)
            plt.close()
        with shap_c2:
            st.pyplot(_plot_shap(sv_inst, feat_labels_inst, "🏛️ RF Institutionnel"), use_container_width=True)
            plt.close()

    # ── Recommandations — 100% dynamiques ─────────────────
    with col_reco:
        st.markdown("#### 💡 Recommandations")
        if est_alerte:
            # Les recommandations sont générées depuis les SHAP values réelles
            recos = generer_recommandations(sv, feat_labels)
            for reco in recos:
                if isinstance(reco, dict):
                    niveau  = reco.get("niveau", "léger")
                    couleur = {"léger": "#c9a84c", "modéré": "#e67e22", "sévère": "#c0392b"}.get(niveau, "#2d6a9f")
                    bg      = {"léger": "#fffbf0", "modéré": "#fff5eb", "sévère": "#fdf0ee"}.get(niveau, "#eef4ff")
                    st.markdown(
                        f'<div class="reco-item" style="background:{bg};border-left-color:{couleur};">'
                        f'<strong style="color:{couleur};">{reco.get("badge","")}</strong> {reco.get("message","")}'
                        f'</div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown(f'<div class="reco-item">{reco}</div>', unsafe_allow_html=True)
        else:
            # Profil sécurisé : messages contextualisés selon le score
            if risque_final < 0.25:
                msg_suivi = "Prochain point de suivi recommandé dans 8 semaines."
            elif risque_final < 0.40:
                msg_suivi = "Prochain point de suivi recommandé dans 4 semaines."
            else:
                msg_suivi = "Prochain point de suivi recommandé dans 2 semaines (score proche du seuil)."
            st.markdown(f'<div class="reco-ok">✅ Profil satisfaisant. Maintenir le suivi de routine.</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="reco-ok">📅 {msg_suivi}</div>', unsafe_allow_html=True)

        # Niveau de confiance
        st.markdown("<br>", unsafe_allow_html=True)
        confiance_model = float(max(risque_final, 1 - risque_final))
        niveau_conf  = "Élevée"  if confiance_model > 0.80 else ("Modérée" if confiance_model > 0.65 else "Faible")
        couleur_conf = "#1e7e5a" if confiance_model > 0.80 else ("#f39c12" if confiance_model > 0.65 else "#c0392b")
        st.markdown(f"""
        <div class="metric-mini" style="margin-top:0.5rem;">
            <div class="metric-mini-val" style="color:{couleur_conf};">{confiance_model:.0%}</div>
            <div class="metric-mini-label">Confiance du modèle — {niveau_conf}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── Sauvegarde historique — saisie manuelle ───────────
    recos_hist = generer_recommandations(sv, feat_labels) if est_alerte else []
    st.session_state.historique.append({
        "type":       "manuel",
        "horodatage": datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
        "statut":     "🔴 ALERTE" if est_alerte else "🟢 OK",
        "risque":     f"{pct}%",
        "p_rf":       f"{p_rf:.1%}",
        "p_inst":     f"{p_inst:.1%}",
        "recos":      recos_hist,
        # snapshot des valeurs clés pour le PDF
        "age": age, "genre": genre, "niveau": niveau,
        "moyenne": moyenne, "absences": absences,
        "stress": stress, "fatigue": fatigue, "motivation": motivation,
    })

    # ── Export PDF de la fiche étudiant ───────────────────
    st.markdown("---")
    st.markdown("#### 📄 Fiche PDF de l'analyse")

    def generer_pdf_fiche(data: dict) -> bytes:
        from fpdf import FPDF

        # ── Fonction de nettoyage : supprime les caractères non-Latin-1
        def clean(text: str) -> str:
            if not isinstance(text, str):
                text = str(text)
            replacements = {
                "\u2014": "-",    # em dash —
                "\u2013": "-",    # en dash –
                "\u2026": "...",  # ellipsis …
                "\u2019": "'",    # apostrophe courbe '
                "\u2018": "'",    # apostrophe courbe '
                "\u201c": '"',    # guillemet "
                "\u201d": '"',    # guillemet "
                "\u00e9": "e",    # é
                "\u00e8": "e",    # è
                "\u00ea": "e",    # ê
                "\u00eb": "e",    # ë
                "\u00e0": "a",    # à
                "\u00e2": "a",    # â
                "\u00ee": "i",    # î
                "\u00ef": "i",    # ï
                "\u00f4": "o",    # ô
                "\u00f9": "u",    # ù
                "\u00fb": "u",    # û
                "\u00fc": "u",    # ü
                "\u00e7": "c",    # ç
                "\u00c9": "E",    # É
                "\u00c0": "A",    # À
                "\u00c7": "C",    # Ç
            }
            for char, repl in replacements.items():
                text = text.replace(char, repl)
            # Supprimer tout caractère restant hors Latin-1 (emojis, etc.)
            return text.encode("latin-1", errors="ignore").decode("latin-1")

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # En-tête
        pdf.set_fill_color(15, 31, 61)          # bleu-nuit
        pdf.rect(0, 0, 210, 32, style='F')
        pdf.set_text_color(201, 168, 76)         # or
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_xy(10, 8)
        pdf.cell(0, 7, clean("SYSTEME D'ALERTE PRECOCE - Fiche d'analyse"), ln=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(200, 200, 200)
        pdf.set_xy(10, 17)
        pdf.cell(0, 6, clean(f"Genere le {data['horodatage']}  |  Modeles : RF comportemental + RF institutionnel  |  Interpretabilite : SHAP"), ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_y(40)

        # Statut global
        alerte = data["statut"] == "🔴 ALERTE"
        if alerte:
            pdf.set_fill_color(253, 240, 238)
            pdf.set_draw_color(192, 57, 43)
        else:
            pdf.set_fill_color(237, 250, 245)
            pdf.set_draw_color(30, 126, 90)
        pdf.set_line_width(0.8)
        pdf.rect(10, pdf.get_y(), 190, 20, style='FD')
        pdf.set_font("Helvetica", "B", 12)
        if alerte:
            pdf.set_text_color(192, 57, 43)
        else:
            pdf.set_text_color(30, 126, 90)
        statut_label = "ALERTE - Risque eleve de decrochage" if alerte else "SECURITE - Risque faible de decrochage"
        pdf.set_xy(14, pdf.get_y() + 4)
        pdf.cell(100, 7, clean(statut_label))
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_xy(160, pdf.get_y() - 4)
        pdf.cell(40, 12, clean(f"Risque : {data['risque']}"), align="R")
        pdf.set_y(pdf.get_y() + 26)
        pdf.set_text_color(0, 0, 0)
        pdf.set_line_width(0.2)

        # Scores détaillés
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(26, 58, 107)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(190, 7, "  Detail des scores", fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 10)
        pdf.set_fill_color(240, 242, 246)
        pdf.cell(95, 8, clean(f"  Score comportemental (RF x60%) : {data['p_rf']}"), fill=True, border=1)
        pdf.cell(95, 8, clean(f"  Score institutionnel (RF x40%) : {data.get('p_inst', '-')}"), fill=True, border=1, ln=True)
        pdf.ln(4)

        # Profil de l'étudiant
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(26, 58, 107)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(190, 7, "  Profil de l'etudiant", fill=True, ln=True)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 10)
        champs = [
            ("Age",        data.get("age",        "-")),
            ("Genre",      data.get("genre",      "-")),
            ("Niveau",     data.get("niveau",     "-")),
            ("Moyenne",    data.get("moyenne",    "-")),
            ("Absences",   data.get("absences",   "-")),
            ("Stress",     data.get("stress",     "-")),
            ("Fatigue",    data.get("fatigue",    "-")),
            ("Motivation", data.get("motivation", "-")),
        ]
        for i, (label, val) in enumerate(champs):
            fill = i % 2 == 0
            if fill:
                pdf.set_fill_color(248, 250, 252)
            else:
                pdf.set_fill_color(255, 255, 255)
            pdf.cell(60, 7, clean(f"  {label}"), fill=True, border=1)
            pdf.cell(130, 7, clean(f"  {val}"), fill=True, border=1, ln=True)
        pdf.ln(4)

        # Recommandations
        if data.get("recos"):
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_fill_color(26, 58, 107)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(190, 7, "  Recommandations", fill=True, ln=True)
            pdf.set_text_color(0, 0, 0)
            couleurs_niv = {"severe": (192,57,43), "modere": (230,126,34), "leger": (201,168,76)}
            for reco in data["recos"]:
                if isinstance(reco, dict):
                    niv   = clean(reco.get("niveau", "leger"))
                    msg   = clean(reco.get("message", ""))
                    badge = clean(reco.get("badge", ""))
                    r,g,b = couleurs_niv.get(niv, (100,100,100))
                    pdf.set_fill_color(r,g,b)
                    pdf.set_text_color(255,255,255)
                    pdf.set_font("Helvetica", "B", 8)
                    pdf.cell(30, 7, f"  {badge[:12]}", fill=True, border=1)
                    pdf.set_fill_color(253,251,245)
                    pdf.set_text_color(0,0,0)
                    pdf.set_font("Helvetica", "", 9)
                    msg_court = msg[:90] + ("..." if len(msg) > 90 else "")
                    pdf.cell(160, 7, f"  {msg_court}", fill=True, border=1, ln=True)
            pdf.ln(3)

        # Footer
        pdf.set_y(-20)
        pdf.set_font("Helvetica", "I", 7)
        pdf.set_text_color(150,150,150)
        pdf.cell(0, 5, "Ce document est un outil d'aide a la decision. Il ne remplace pas le jugement pedagogique humain.", align="C")

        return bytes(pdf.output())

    col_pdf1, col_pdf2, col_pdf3 = st.columns([1, 2, 1])
    with col_pdf2:
        pdf_data = generer_pdf_fiche(st.session_state.historique[-1])
        st.download_button(
            label="📄  Télécharger la fiche PDF",
            data=pdf_data,
            file_name=f"fiche_alerte_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

# ─────────────────────────────────────────────────────────────
# HISTORIQUE DES ANALYSES (session courante)
# ─────────────────────────────────────────────────────────────
if st.session_state.historique:
    st.markdown("---")
    with st.expander(f"🕐 Historique de la session — {len(st.session_state.historique)} analyse(s)", expanded=False):
        col_hist_btn = st.columns([3, 1])[1]
        with col_hist_btn:
            if st.button("🗑️ Vider l'historique", use_container_width=True):
                st.session_state.historique = []
                st.rerun()

        for i, entree in enumerate(reversed(st.session_state.historique)):
            idx = len(st.session_state.historique) - i
            if entree["type"] == "manuel":
                couleur_hist = "#fdf0ee" if entree["statut"] == "🔴 ALERTE" else "#edfaf5"
                bordure_hist = "#c0392b" if entree["statut"] == "🔴 ALERTE" else "#1e7e5a"
                st.markdown(
                    f'<div style="background:{couleur_hist};border-left:4px solid {bordure_hist};'
                    f'border-radius:0 8px 8px 0;padding:0.6rem 1rem;margin-bottom:0.5rem;">'
                    f'<strong>#{idx}</strong> · {entree["horodatage"]} · Saisie manuelle · '
                    f'{entree["statut"]} · Risque : <strong>{entree["risque"]}</strong> '
                    f'(RF {entree["p_rf"]} / RF inst. {entree.get("p_inst", "-")})'
                    f'</div>',
                    unsafe_allow_html=True
                )
            elif entree["type"] == "batch":
                st.markdown(
                    f'<div style="background:#f0f4ff;border-left:4px solid #2d6a9f;'
                    f'border-radius:0 8px 8px 0;padding:0.6rem 1rem;margin-bottom:0.5rem;">'
                    f'<strong>#{idx}</strong> · {entree["horodatage"]} · Import batch · '
                    f'📁 {entree["fichier"]} · '
                    f'{entree["nb_total"]} étudiant(s) · '
                    f'🔴 {entree["nb_alerte"]} alerte(s)'
                    f'</div>',
                    unsafe_allow_html=True
                )
                with st.expander(f"Voir le tableau — analyse #{idx}", expanded=False):
                    st.dataframe(entree["df_res"], use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
    Système d'Alerte Précoce — Décrochage Académique · Modèles : RF comportemental + RF institutionnel · Interprétabilité : SHAP<br>
    <em>Ce système est un outil d'aide à la décision. Il ne remplace pas le jugement pédagogique humain.</em>
</div>
""", unsafe_allow_html=True)